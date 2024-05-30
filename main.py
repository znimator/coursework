from PyQt6.QtCore import *
from PyQt6.QtWidgets import *
import openpyxl
import sys
import time

# Изменение метода __lt__ чтобы можно было производить поиск текстом
class TableWidgetItem(QTableWidgetItem):
    def __lt__(self, other):
        try:
            return float(self.text()) < float(other.text())
        except ValueError:
            return super().__lt__(other)

# Основное окно приложения
class Window(QMainWindow):
    def __init__(self):
        super().__init__()

        self.menu_bar = QMenuBar()

        self.menu = QMenu("Файл")
        action1 = self.menu.addAction("Сохранить")
        action1.triggered.connect(self.export)
        self.setStyleSheet("background-color: #424242; color: black;")
        

        self.menu_bar.addMenu(self.menu)

        self.setMenuBar(self.menu_bar)

    def closeEvent(self, event):
        print("closing")
        widget.DataWindow.close()

        event.accept()

    def export(self):
        filename, filter = QFileDialog.getSaveFileName(self, 'Save file', '','Excel files (*.xlsx)')
        wb = openpyxl.Workbook()

        temp = wb["Sheet"]
        wb.remove(temp)

        print(filename)

        centralWidget = self.centralWidget()

        for sheet, table_widget in centralWidget.temp_sheets.items():
            workingSheet = wb.create_sheet(sheet.title)

            labels = []
            for c in range(table_widget.columnCount()):
                it = table_widget.horizontalHeaderItem(c)
                labels.append(str(c+1) if it is None else it.text())
            print(labels)

            for i in range(len(labels)):
                value = labels[i]
                workingSheet.cell(1, i + 1, value)

            if filename:
                for column in range(table_widget.columnCount()):
                    for row in range(table_widget.rowCount()):
                        try:
                            text = str(table_widget.item(row, column).text())
                            workingSheet.cell(row + 2, column + 1, text)
                        except AttributeError:
                            pass
                        
        wb.save(filename)

# Окно для добавления новой даты
class DataWindow(QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Add Data")
        #self.setFixedSize(500, 500)
        self.setStyleSheet("background-color: #424242; color: white;")

        self.Layout = QVBoxLayout()
        self.setLayout(self.Layout)

        self.addButton = QPushButton("Добавить")
        self.cancelButton = QPushButton("Отмена")
        self.cancelButton.clicked.connect(self.cancel)
        self.addButton.clicked.connect(self.acceptInfo)
        self.Layout.addWidget(self.addButton)
        
        self.inputs :list[QLineEdit] = []

    def cancel(self):
        self.close()

    # Добавления новой строки по записанным данным
    def acceptInfo(self):
        widget.currentWidget.insertRow(0)

        for i in range(len(self.inputs)):
            print(0, i)
            item = TableWidgetItem(str(self.inputs[i].text()))

            widget.currentWidget.setItem(0 , i, item)

    #Загрузка полей для заполнений
    def load(self):
        print("loading stuff")

        self.Layout.removeWidget(self.addButton)

        labels = []

        for c in range(widget.currentWidget.columnCount()):
            label = widget.currentWidget.horizontalHeaderItem(c)
            labels.append(str(c+1) if label is None else label.text())
        
        for l in range(len(labels)):
            lineEdit = QLineEdit()
            lineEdit.setPlaceholderText(labels[l])
            self.Layout.addWidget(lineEdit)

            self.inputs.append(lineEdit)

        self.Layout.addWidget(self.addButton)

    # Действия при закрытии окна
    def closeEvent(self, event):
        print("clearing stuff")

        for i in reversed(range(self.Layout.count())): 
            self.Layout.itemAt(i).widget().setParent(None)

        self.inputs :list[QLineEdit] = []

        event.accept()

# Главное окно где создаются списки и редактируется дата
class Main(QWidget):
    def __init__(self):
        super(Main, self).__init__()
        self.setWindowTitle("Load Excel data to QTableWidget")
        self.setBaseSize(700, 700)
        
        self.main_layout = QVBoxLayout()
        self.setLayout(self.main_layout)
        
        self.addData = QPushButton("Добавить")
        self.remove_row = QPushButton("Удалить линию")

        self.mainLabel = QLabel("МАСТЕР, ВЛАДЫКА")
    
        self.filter_list = QComboBox()

        self.DataWindow = DataWindow()

        self.addData.clicked.connect(self.show_data_window)

        self.remove_row.clicked.connect(self.deleteRow)

        self.selectedRow = 0

        self.search = QLineEdit()
        self.search.setPlaceholderText("Поиск по Фильтру")
        self.search.textChanged.connect(self.findName)

        #Добавление виджетов pyqt6 в макет/каркас
        self.main_layout.addWidget(self.mainLabel, alignment=Qt.AlignmentFlag.AlignCenter)
        self.main_layout.addWidget(self.filter_list)
        self.main_layout.addWidget(self.search)
        self.main_layout.addWidget(self.addData)
        self.main_layout.addWidget(self.remove_row)

        
        path = "./data.xlsx"
        self.workbook = openpyxl.load_workbook(path)
        self.temp_sheets = {}

        self.buttons :dict[str, QPushButton] = {}

        self.expandButtons = []

        # Предзагрузка листов
        for name in self.workbook.sheetnames:
            tableWidget = QTableWidget()
            label = QLabel(name)
            expand = QPushButton("Развернуть")
            expand.linked = tableWidget
            expand.linkedLabel = label

            expand.clicked.connect(self.expandTalbe)
            self.expandButtons.append(expand)

            tableWidget.cellClicked.connect(self.cellClicked)
            self.temp_sheets[self.workbook[name]] = tableWidget

            self.main_layout.addWidget(label)
            self.main_layout.addWidget(expand)
            self.main_layout.addWidget(tableWidget)

        self.currentWidget :QTableWidget = self.temp_sheets[next(iter(self.temp_sheets))]

        self.load_data()

        filter_labels = []
        print(self.temp_sheets)
        for sheetNum in range(len(self.temp_sheets)):
            sheet = self.temp_sheets[self.workbook.worksheets[sheetNum]]
            print(sheet)
            print(sheet.columnCount())
            for c in range(sheet.columnCount()):
                label = sheet.horizontalHeaderItem(c)
                toAppend = str(c+1) if label is None else label.text()
                filter_labels.append(toAppend)

        print(filter_labels)

        for i in filter_labels:
            self.filter_list.addItem(i)  

    def expandTalbe(self):
        sender = self.sender()
        print(sender.linked)
        if sender.text() == "Развернуть":
            sender.setText("Свернуть")
            for button in self.expandButtons:
                if button != sender:
                    button.hide()
                    button.linked.hide()
                    button.linkedLabel.hide()
        else:
            for button in self.expandButtons:
                button.setText("Развернуть")
                button.show()
                button.linked.show()
                button.linkedLabel.show()

    #Запуск окна с добавлением инфы
    def show_data_window(self):
        self.DataWindow.show()
        self.DataWindow.load()

    #Сортировка
    def findName(self):
        name = self.search.text().lower()
        found = False

        #Поиск по заданной строке
        for sheetNum in range(len(self.temp_sheets)):
            sheet = self.temp_sheets[self.workbook.worksheets[sheetNum]]

            for row in range(sheet.rowCount()):
                for column in range(sheet.columnCount()):
                    header = sheet.horizontalHeaderItem(column)
                    if header != None:
                        if header.text() == self.filter_list.currentText(): 
                            item = sheet.item(row, column)
                            try:
                                found = name in item.text().lower()
                                sheet.setRowHidden(row, not found)
                                if found:
                                    break
                            except AttributeError:
                                pass

    #Переключение между листами Excel при помощи QListWidget
    def listClick(self):
        sender = self.sender()
        print(sender)
        item = sender

        for sheet, widget in self.temp_sheets.items():
            widget.hide()

        self.temp_sheets[self.workbook[item.text()]].show()
        self.currentWidget = self.temp_sheets[self.workbook[item.text()]]

        labels = []

        for c in range(self.currentWidget.columnCount()):
            label = self.currentWidget.horizontalHeaderItem(c)
            labels.append(str(c+1) if label is None else label.text())
    
        #Загрузка фильтра при смене листа
        self.filter_list.clear()

        for i in labels:
            self.filter_list.addItem(i)

    #Загрузка всех листов Excel при помощи .loadSheet()
    def load_data(self):

        for name in self.workbook.sheetnames:
            self.loadSheet(self.workbook[name])

    #Загрузить лист Excel
    def loadSheet(self, sheet):

        self.temp_sheets[sheet].setRowCount(sheet.max_row)
        self.temp_sheets[sheet].setColumnCount(sheet.max_column)
        
        list_values = list(sheet.values)
        self.temp_sheets[sheet].setHorizontalHeaderLabels(list_values[0])
        self.temp_sheets[sheet].setSortingEnabled(True)

        row_index = 0

        # Добавление информации из листа Excel в QTableWidget
        for value_tuple in list_values[1:]:
            col_index = 0
            for value in value_tuple:
                item = TableWidgetItem(str(value))
                
                self.temp_sheets[sheet].setItem(row_index , col_index, item)
               
                col_index += 1
            row_index += 1

    #Добавить линию
    def addRow(self):
        self.currentWidget.insertRow(self.selectedRow)

        return self.selectedRow

    #Удалить линию
    def deleteRow(self, *args):
        print(self.selectedRow)
        self.currentWidget.removeRow(self.selectedRow)
        print(*args)
    
    # Сохранить текущую клетку
    def cellClicked(self, row, column):
        print(self.sender())
        
        self.selectedRow = row

        self.currentWidget = self.sender()
        

#Запуск кода
if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = Window()
    widget = Main()

    window.setCentralWidget(widget)
    widget.show()
    window.show()
    
    app.exec()